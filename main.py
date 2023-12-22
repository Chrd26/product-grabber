import urllib.robotparser as urobot
import requests
from bs4 import BeautifulSoup
import sys
import os
from urllib.parse import urlparse
import pandas as pd
from openpyxl import load_workbook

startrow = 0

if os.path.exists('output.xlsx'):
    os.remove('output.xlsx')

# CLI arguments
if len(sys.argv) == 1:
    print("Not enough arguments. You need to also add a url to run the tool\n")
    print("Example: python3 main.py www.example.com\n")
    exit(1)


for i in range(1, len(sys.argv)):

    # Create a RobotFileParser object and set it's URL to example.com/robots.txt
    rp = urobot.RobotFileParser()
    rp.set_url(sys.argv[i])
    rp.read()

    # Check if the website allows web scraping
    if rp.can_fetch('*', sys.argv[1]):
        # Make a request to the website
        res = requests.get(sys.argv[1])
        res.raise_for_status()

        # Parse the HTML content
        soup = BeautifulSoup(res.text, 'html.parser')

        # Find the images you want to extract
        img_elements = soup.find_all('img')

        for i, img in enumerate(img_elements):
            # Get the source URL of the image
            img_url = img.get('src')

            # Skip if the url is not absolute
            if not img_url.startswith('http'):
                continue

            # Make a request to the image URL
            img_res = requests.get(img_url)

            # Extract the original file name from the URL
            parsed_url = urlparse(img_url)
            filename = os.path.basename(parsed_url.path)

            # Write the image to a file
            with open(filename, 'wb') as f:
                f.write(img_res.content)

        # Extract text
        collection = ['span.custom-sku', 'h1.page-title',
                      'span.price', 'div.value']
        text_elements = soup.select(', '.join(collection))
        text_content = [element.text for element in text_elements]

        # Write text to Excel
        df = pd.DataFrame(text_content, columns=['Text'])

        if os.path.exists('output.xlsx'):
            book = load_workbook('output.xlsx')
            writer = pd.ExcelWriter('output.xlsx', engine='openpyxl', mode='a',
                                    if_sheet_exists='overlay')
            # writer.book = book

            # Check for existing sheet
            if 'Sheet1' in book.sheetnames:
                print("Hello")
                startrow = writer.sheets['Sheet1'].max_row
                df.to_excel(writer, index=False, header=False,
                            startrow=startrow, sheet_name='Sheet1')
            else:
                df.to_excel(writer, index=False, startrow=startrow, 
                            sheet_name='Sheet1')

            startrow += 2
            startrow += len(df)
            # Handle tables
            # comment out when it is not needed
            table = pd.read_html(res.text)
            for i, table in enumerate(table):
                table.to_excel(writer, index=False, header=False,
                               startrow=startrow, sheet_name='Sheet1')
                startrow += len(table)

            writer._save()
        else:
            writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
            df.to_excel(writer, index=False, startrow=startrow,
                        sheet_name='Sheet1')

            startrow += 3
            startrow += len(df)
            # Handle tables
            # comment out when it is not needed
            table = pd.read_html(res.text)
            for i, table in enumerate(table):
                table.to_excel(writer, index=False, header=False,
                               startrow=startrow, sheet_name='Sheet1')
                startrow += len(table)
            writer._save()

        writer.close()
    else:
        print("This website does not allow web scraping.")
